# -*-coding: UTF-8 -*-
import sys
import time
from mylogclass import MyLogClass
from PyQt5.QtWidgets import QApplication, QFrame,QFileDialog
from UI.UI_win import Ui_Form
from PyQt5.QtCore import Qt
import openpyxl
import docx
import re
from BaiduTranslate.BaiduTranslateSpider import BaiduTranslateSpider
from threading import Thread


class MainWindow(QFrame, Ui_Form):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)

        self.setupUi(self)
        self.retranslateUi(self)

        self.setWindowFlags(
            Qt.Window | Qt.FramelessWindowHint | Qt.WindowSystemMenuHint | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint)  # 窗口无边框

        self.pushButton_3.clicked.connect(self.close)
        self.pushButton.clicked.connect(self.open_file)
        self.pushButton_4.clicked.connect(self.open_excel)
        self.pushButton_2.clicked.connect(self.out_excel)
        self.file_path_list = []
        self.file_path_excel = None
        self.thread_count = 20
        self.result = []
        self.baidu = BaiduTranslateSpider()
        self.log = MyLogClass()
    def open_excel(self):
        self.file_path_excel,_ = QFileDialog.getOpenFileName(self, '选择文件', '', 'Excel files(*.xlsx;)')
        self.lineEdit.setText(self.file_path_excel)
    def open_file(self):
        self.file_path_list,_ = QFileDialog.getOpenFileNames(self, '选择文件', '', 'Word files(*.docx;)')
        self.label_2.setText(str(len(self.file_path_list)))


    def exWord(self,word):
        reuslt = self.baidu.get_result(word)
        self.log.logger.info(reuslt)
        print(reuslt)
        self.result.append(reuslt)


    # def exWord(self,wordList):
    #     baidu = BaiduTranslateSpider()
    #     result = []
    #     for word in wordList:
    #         result.append(baidu.get_result(word))
    #     return result
    def thread_pool(self,wordList):
        for i in range(0, len(wordList), self.thread_count):
            task_list = [Thread(target=self.exWord, args=(word,)) for word in wordList[i:i + self.thread_count]]
            for task in task_list:
                time.sleep(0.5)
                task.start()
            [task.join() for task in task_list]


    def out_excel(self):
        if self.file_path_list:
            textList = ''
            result_list = []
            for doc in self.file_path_list:
                word = docx.Document(doc)
                for para in word.paragraphs:
                    textList += para.text

            if self.file_path_excel:
                wb = openpyxl.load_workbook(self.file_path_excel)
                ws = wb.get_active_sheet()
                # wordList = set([a.value for a in  ws['A']])
                # wordList = self.exWord(wordList)

                wordList = [a.value for a in  ws['A']]
                self.thread_pool(wordList)


                for word in self.result:
                    for k,v in word.items():
                        count = 0
                        for i in v:
                            result = re.findall('(^|[^a-zA-Z])'+i+'($|[^a-zA-Z])',textList,flags=re.S|re.I)
                            count += len(result)
                        result_list.append([k,count])

            if result_list:
                result_list = sorted(result_list, key=lambda x: x[1], reverse=True)
                fileName, ok = QFileDialog.getSaveFileName(None, "文件保存", "./", "Excel (*.xlsx)")
                workbook = openpyxl.Workbook()
                booksheet = workbook.active
                booksheet.append(['单词', '次数'])
                [booksheet.append(li) for li in result_list]
                workbook.save(fileName)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.m_flag = True
            self.m_Position = event.globalPos() - self.pos()  # 获取鼠标相对窗口的位置
            event.accept()

    def mouseMoveEvent(self, QMouseEvent):
        if Qt.LeftButton and self.m_flag:
            self.move(QMouseEvent.globalPos() - self.m_Position)  # 更改窗口位置
            QMouseEvent.accept()

    def mouseReleaseEvent(self, QMouseEvent):
        self.m_flag = False

if __name__ == '__main__':
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())
